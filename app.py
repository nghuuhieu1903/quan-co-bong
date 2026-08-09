from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_session import Session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import qrcode
import io
import base64
import json
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import threading
import time
try:
    import pyttsx3
except ImportError:
    pyttsx3 = None

try:
    import pyautogui
except ImportError:
    pyautogui = None

try:
    import winsound
except ImportError:
    winsound = None

try:
    import gtts
    from gtts import gTTS
except ImportError:
    gtts = None
    gTTS = None

try:
    import pygame  # type: ignore
except ImportError:
    pygame = None

import tempfile
import re
from sqlalchemy import text, func
import builtins
import sys
from functools import wraps

try:
    from dotenv import load_dotenv
    # Loads variables from a .env file placed next to app.py (if present).
    # Lets DATABASE_URL/SECRET_KEY be set via a plain file on the VPS
    # instead of relying on a control panel's "environment variables" UI.
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))
except ImportError:
    pass

def safe_print(*args, **kwargs):
    safe_args = []
    encoding = sys.stdout.encoding or 'utf-8'
    for arg in args:
        if isinstance(arg, str):
            try:
                arg.encode(encoding)
                safe_args.append(arg)
            except UnicodeEncodeError:
                safe_args.append(arg.encode(encoding, errors='replace').decode(encoding, errors='replace'))
        else:
            safe_args.append(arg)
    builtins.print(*safe_args, **kwargs)

print = safe_print

app = Flask(__name__)

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')

# MySQL database configuration.
# Set DATABASE_URL to e.g. mysql+pymysql://user:password@host:3306/dbname
# Falls back to a local MySQL default for development if not set.
db_url = os.environ.get('DATABASE_URL', 'mysql+pymysql://root:@localhost:3306/ecommerce')
if db_url.startswith('mysql://'):
    db_url = db_url.replace('mysql://', 'mysql+pymysql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
# Force utf8mb4 on every connection so Vietnamese text is stored correctly
# regardless of the MySQL server's default charset.
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'connect_args': {'charset': 'utf8mb4'}}

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_TYPE'] = 'filesystem'
app.config['UPLOAD_FOLDER'] = 'static/images'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Laptop Speaker Configuration
class LaptopSpeaker:
    def __init__(self):
        self.engine = None  # pyttsx3 engine (fallback)
        self.enabled = True  # Enable/disable announcements
        self.voice_rate = 150  # Speech rate (for pyttsx3)
        self.voice_volume = 0.9  # Volume level (for pyttsx3)
        self.voice_model = None  # Voice model for Vietnamese
        self.use_gtts = True  # Use gTTS for Vietnamese by default
        self.tts_language = 'vi'  # Language for gTTS
        
        # Initialize pygame mixer for gTTS audio playback
        if pygame is not None:
            try:
                pygame.mixer.init()
                print("Pygame mixer initialized successfully")
            except Exception as e:
                print(f"Error initializing pygame mixer: {e}")
                self.use_gtts = False
        else:
            print("Pygame not installed, disabling gTTS audio playback")
            self.use_gtts = False
        
    def initialize_engine(self):
        """Initialize text-to-speech engine with fallback options"""
        try:
            if self.use_gtts:
                # For gTTS, we don't need to initialize engine
                print("Using gTTS for Vietnamese text-to-speech")
                return True
            else:
                # Fallback to pyttsx3
                if pyttsx3 is None:
                    print("pyttsx3 is not installed, cannot use fallback")
                    return False
                self.engine = pyttsx3.init()
                
                # Lấy danh sách các voice có sẵn
                voices = self.engine.getProperty('voices')
                
                # Tìm voice tiếng Việt (Microsoft An, Microsoft HoaiMy, hoặc các voice khác)
                vietnamese_voice = None
                for voice in voices:
                    voice_id_lower = voice.id.lower() if voice.id else ''
                    voice_name_lower = voice.name.lower() if voice.name else ''
                    # Tìm chính xác hơn: ưu tiên Microsoft An và Microsoft HoaiMy
                    if ('microsoft an' in voice_name_lower or 'microsoft hoaimy' in voice_name_lower or
                        'vi_' in voice_id_lower or 'vietnamese' in voice_name_lower or 
                        'vietnam' in voice_name_lower or voice_name_lower.startswith('vi ')):
                        vietnamese_voice = voice
                        break
                
                # Nếu không tìm thấy voice tiếng Việt, sử dụng voice đầu tiên
                if vietnamese_voice:
                    self.engine.setProperty('voice', vietnamese_voice.id)
                    print(f"Using Vietnamese voice: {vietnamese_voice.name}")
                else:
                    # Sử dụng voice mặc định và hiển thị thông báo
                    default_voice = voices[0] if voices else None
                    if default_voice:
                        self.engine.setProperty('voice', default_voice.id)
                        print(f"Using default voice: {default_voice.name}")
                    print("Warning: No Vietnamese voice found. Using default voice.")
                    print("Available voices:")
                    for i, voice in enumerate(voices):
                        print(f"  {i+1}. {voice.name} (ID: {voice.id})")
                
                # Áp dụng cài đặt
                self.engine.setProperty('rate', self.voice_rate)
                self.engine.setProperty('volume', self.voice_volume)
                
                return True
        except Exception as e:
            print(f"Error initializing TTS engine: {e}")
            return False
    
    def speak_with_gtts(self, text, lang='vi'):
        """Speak text using Google Text-to-Speech (gTTS)"""
        if gTTS is None or pygame is None:
            print("gTTS or pygame is not installed. Trying pyttsx3 fallback if available.")
            if self.engine:
                try:
                    self.engine.say(text)
                    self.engine.runAndWait()
                except Exception as fallback_error:
                    print(f"Fallback TTS also failed: {fallback_error}")
            return
        try:
            print(f"Generating speech with gTTS: {text}")
            
            # Create gTTS object
            tts = gTTS(text=text, lang=lang, slow=False)
            
            # Save to temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.mp3') as temp_file:
                temp_filename = temp_file.name
                tts.save(temp_filename)
            
            # Play the audio file
            try:
                pygame.mixer.music.load(temp_filename)
                pygame.mixer.music.play()
                
                # Wait for the audio to finish playing
                while pygame.mixer.music.get_busy():
                    time.sleep(0.1)
                
                print("gTTS speech completed successfully")
            except Exception as e:
                print(f"Error playing gTTS audio: {e}")
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_filename)
                except:
                    pass
                    
        except Exception as e:
            print(f"Error with gTTS: {e}")
            # Fallback to pyttsx3 if available
            if self.engine:
                try:
                    self.engine.say(text)
                    self.engine.runAndWait()
                except Exception as fallback_error:
                    print(f"Fallback TTS also failed: {fallback_error}")
    
    def play_notification_sound(self):
        """Play a simple notification sound"""
        if winsound is not None:
            try:
                # Play Windows notification sound
                winsound.MessageBeep(winsound.MB_OK)
            except Exception as e:
                print(f"Error playing notification sound: {e}")
        else:
            print("winsound is not available")
    
    def announce_order(self, order_items, customer_name="Khách", notes=None):
        """Announce order details through laptop speakers"""
        if not self.enabled:
            print("Speaker is disabled, skipping announcement")
            return
        
        # Prepare product details before starting thread (to avoid database context issues)
        product_details = []
        for item in order_items:
            try:
                product = Product.query.get(item['product_id'])
                if product:
                    product_details.append(f"{product.name}, số lượng {item['quantity']}")
                else:
                    product_details.append(f"Sản phẩm {item['product_id']}, số lượng {item['quantity']}")
            except Exception as e:
                print(f"Error getting product {item['product_id']}: {e}")
                product_details.append(f"Sản phẩm {item['product_id']}, số lượng {item['quantity']}")
        
        def speak_order():
            try:
                print("Starting order announcement...")
                # Play notification sound first
                self.play_notification_sound()
                
                # Prepare announcement message in Vietnamese
                message = f"Đơn hàng mới! {customer_name} đã đặt hàng: "
                message += ". ".join(product_details)
                message += ". Xin vui lòng chuẩn bị đơn hàng!"

                if notes:
                    message += f". Ghi chú: {notes}"
                
                print(f"Announcing: {message}")
                
                # Use gTTS for Vietnamese (better quality)
                if self.use_gtts:
                    self.speak_with_gtts(message, lang='vi')
                else:
                    # Fallback to pyttsx3
                    if not self.ensure_engine_ready():
                        print("Cannot initialize TTS engine for order announcement")
                        return
                    print(f"Engine status before speaking: {'ready' if self.engine else 'None'}")
                    self.engine.say(message)
                    self.engine.runAndWait()
                
                print("Order announcement completed successfully")
                
            except Exception as e:
                print(f"Error in speech announcement: {e}")
                # Thử khởi tạo lại engine cho lần sau
                self.engine = None
        
        # Run speech in separate thread to avoid blocking
        speech_thread = threading.Thread(target=speak_order)
        speech_thread.daemon = True
        speech_thread.start()
    
    def test_speaker(self):
        """Test the laptop speaker"""
        def speak_test():
            try:
                # Play test sound
                self.play_notification_sound()
                
                message = "Kiểm tra loa laptop. Hệ thống đã sẵn sàng."
                
                # Use gTTS if enabled, otherwise fallback to pyttsx3
                if self.use_gtts:
                    print("Testing with gTTS...")
                    self.speak_with_gtts(message, lang='vi')
                    return True
                else:
                    print("Testing with pyttsx3...")
                    # Đảm bảo engine sẵn sàng
                    if not self.ensure_engine_ready():
                        print("Cannot initialize TTS engine for test")
                        return False
                    
                    try:
                        self.engine.say(message)
                        self.engine.runAndWait()
                        return True
                    except Exception as engine_error:
                        print(f"pyttsx3 engine error: {engine_error}")
                        # Try to reinitialize engine
                        self.engine = None
                        if self.initialize_engine():
                            self.engine.say(message)
                            self.engine.runAndWait()
                            return True
                        return False
                    
            except Exception as e:
                print(f"Error testing speaker: {e}")
                # Thử khởi tạo lại engine cho lần sau
                self.engine = None
                return False
        
        test_thread = threading.Thread(target=speak_test)
        test_thread.daemon = True
        test_thread.start()
        return True
    
    def set_voice_settings(self, rate=None, volume=None):
        """Adjust voice settings"""
        print(f"Setting voice settings - rate: {rate}, volume: {volume}")
        print(f"Current engine status: {'exists' if self.engine else 'None'}")
        
        # Cập nhật giá trị nội bộ trước
        if rate is not None:
            self.voice_rate = rate
        if volume is not None:
            self.voice_volume = volume
        
        # Dừng engine hiện tại nếu có
        if self.engine:
            try:
                self.engine.stop()
            except:
                pass
        
        # Khởi tạo lại engine với cài đặt mới
        print("Re-initializing engine with new settings...")
        self.engine = None
        
        if self.initialize_engine():
            print("Engine re-initialized successfully")
            # Áp dụng cài đặt cho engine mới
            try:
                if rate is not None:
                    self.engine.setProperty('rate', rate)
                    print(f"Rate set to: {rate}")
                if volume is not None:
                    self.engine.setProperty('volume', volume)
                    print(f"Volume set to: {volume}")
                print(f"Voice settings updated successfully: rate={rate}, volume={volume}")
            except Exception as e:
                print(f"Error applying settings to engine: {e}")
                # Thử lại một lần nữa
                self.engine = None
                if self.initialize_engine():
                    if rate is not None:
                        self.engine.setProperty('rate', rate)
                    if volume is not None:
                        self.engine.setProperty('volume', volume)
                    print("Settings applied after retry")
        else:
            print("Failed to re-initialize engine")
    
    def get_available_voices(self):
        """Get list of available voices"""
        try:
            if not self.engine:
                self.initialize_engine()
            
            voices = self.engine.getProperty('voices')
            voice_list = []
            
            for i, voice in enumerate(voices):
                voice_info = {
                    'id': voice.id,
                    'name': voice.name,
                    'languages': voice.languages,
                    'gender': voice.gender,
                    'index': i
                }
                voice_list.append(voice_info)
            
            return voice_list
        except Exception as e:
            print(f"Error getting voices: {e}")
            return []
    
    def set_voice_model(self, voice_id):
        """Set voice model by ID"""
        try:
            if not self.engine:
                if not self.initialize_engine():
                    return False
            
            voices = self.engine.getProperty('voices')
            for voice in voices:
                if voice.id == voice_id:
                    self.engine.setProperty('voice', voice.id)
                    self.voice_model = voice_id
                    print(f"Voice model set to: {voice.name}")
                    return True
            
            print(f"Voice ID {voice_id} not found")
            return False
        except Exception as e:
            print(f"Error setting voice model: {e}")
            return False
    
    def ensure_engine_ready(self):
        """Đảm bảo engine sẵn sàng để sử dụng"""
        if not self.engine:
            return self.initialize_engine()
        return True
    
    def toggle_enabled(self):
        """Enable/disable announcements"""
        self.enabled = not self.enabled
        return self.enabled

# Initialize Laptop Speaker
laptop_speaker = LaptopSpeaker()

# PyAutoGUI Automation Class
class AutomationController:
    def __init__(self):
        self.enabled = True
        self.auto_screenshot = True
        self.auto_minimize = False
        self.notification_position = "top-right"
        
        # Set pyautogui settings
        if pyautogui is not None:
            pyautogui.FAILSAFE = True  # Move mouse to corner to stop
            pyautogui.PAUSE = 0.5      # Pause between actions
        else:
            self.enabled = False
    
    def take_order_screenshot(self, order_id):
        """Take screenshot when order is placed"""
        if not self.auto_screenshot or pyautogui is None:
            return
            
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"order_{order_id}_{timestamp}.png"
            
            # Create screenshots directory if it doesn't exist
            screenshot_dir = "screenshots"
            if not os.path.exists(screenshot_dir):
                os.makedirs(screenshot_dir)
            
            filepath = os.path.join(screenshot_dir, filename)
            screenshot = pyautogui.screenshot()
            screenshot.save(filepath)
            
            print(f"Screenshot saved: {filepath}")
            return filepath
        except Exception as e:
            print(f"Error taking screenshot: {e}")
            return None
    
    def show_order_notification(self, order_id, customer_name, total_amount):
        """Show desktop notification for new order"""
        if not self.enabled:
            return
            
        def show_notification():
            try:
                # In ra console
                print(f"THONG BAO DON HANG MOI!")
                print(f"Ma don: #{order_id}")
                safe_customer_name = customer_name.encode('ascii', 'ignore').decode('ascii') if customer_name else 'Khách'
                print(f"Khách hàng: {safe_customer_name}")
                print(f"Tổng tiền: {total_amount:,.0f} VND")
                print("Vui long kiem tra he thong de xu ly don hang.")
                
                # Phát âm thanh thông báo
                if winsound is not None:
                    try:
                        winsound.MessageBeep(winsound.MB_OK)  # Phát âm thanh thông báo
                        time.sleep(0.5)
                        winsound.MessageBeep(winsound.MB_OK)  # Phát lần thứ hai để nhấn mạnh
                    except Exception as e:
                        print(f"Khong the phat am thanh thong bao: {e}")
                else:
                    print("winsound is not available")
                
                # Thêm desktop notification thực sự
                try:
                    # Thử mở cửa sổ thông báo đơn giản với webbrowser
                    import webbrowser

                    # Tạo HTML notification tạm thời
                    html_content = f'''
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <title>Đơn hàng mới!</title>
                        <meta charset="UTF-8">
                        <style>
                            body {{
                                font-family: Arial, sans-serif;
                                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                                color: white;
                                text-align: center;
                                padding: 20px;
                                margin: 0;
                            }}
                            .container {{
                                max-width: 400px;
                                margin: 50px auto;
                                background: rgba(255,255,255,0.1);
                                padding: 30px;
                                border-radius: 15px;
                                backdrop-filter: blur(10px);
                                box-shadow: 0 8px 32px rgba(0,0,0,0.3);
                            }}
                            h1 {{
                                color: #FFD700;
                                margin-bottom: 20px;
                                font-size: 28px;
                            }}
                            .info {{
                                background: rgba(255,255,255,0.2);
                                padding: 15px;
                                border-radius: 10px;
                                margin: 10px 0;
                                font-size: 16px;
                            }}
                            .close-btn {{
                                background: #FF6B6B;
                                color: white;
                                border: none;
                                padding: 10px 20px;
                                border-radius: 5px;
                                cursor: pointer;
                                font-size: 16px;
                                margin-top: 20px;
                            }}
                            .close-btn:hover {{
                                background: #FF5252;
                            }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <h1>🔔 ĐƠN HÀNG MỚI!</h1>
                            <div class="info">
                                <strong>Mã đơn:</strong> #{order_id}<br>
                                <strong>Khách hàng:</strong> {safe_customer_name}<br>
                                <strong>Tổng tiền:</strong> {total_amount:,.0f} VNĐ
                            </div>
                            <p style="margin: 20px 0;">
                                Vui lòng kiểm tra hệ thống để xử lý đơn hàng!
                            </p>
                            <button class="close-btn" onclick="window.close()">Đóng</button>
                        </div>
                        <script>
                            // Tự động đóng sau 30 giây
                            setTimeout(function(){{
                                window.close();
                            }}, 30000);
                            
                            // Phát âm thanh nếu có thể
                            try {{
                                const audio = new Audio('data:audio/wav;base64,UklGRnoGAABXQVZFZm10IBAAAAABAAEAQB8AAEAfAAABAAgAZGF0YQoGAACBhYqFbF1fdJivrJBhNjVgodDbq2EcBj+a2/LDciUFLIHO8tiJNwgZaLvt559NEAxQp+PwtmMcBjiR1/LMeSwFJHfH8N2QQAoUXrTp66hVFApGn+DyvmwhBSuBzvLZiTYIG2m98OScTgwOUarm7blmGgU7k9n1unEiBC13yO/eizEIHWq+8+OWT');
                                audio.play();
                            }} catch(e) {{}}
                        </script>
                    </body>
                    </html>
                    '''
                    
                    # Lưu file HTML tạm
                    with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                        f.write(html_content)
                        temp_file = f.name
                    
                    # Mở trình duyệt với file HTML
                    webbrowser.open(f'file://{temp_file}')
                    print("HTML notification opened in browser")
                    
                    # Xóa file sau 30 giây
                    def cleanup_temp_file():
                        time.sleep(30)
                        try:
                            os.unlink(temp_file)
                        except:
                            pass
                    
                    cleanup_thread = threading.Thread(target=cleanup_temp_file)
                    cleanup_thread.daemon = True
                    cleanup_thread.start()
                    
                except Exception as e:
                    print(f"HTML notification error: {e}")
                    
                    # Fallback: Thử messagebox đơn giản
                    try:
                        import tkinter as tk
                        from tkinter import messagebox

                        def show_messagebox():
                            root = tk.Tk()
                            root.withdraw()
                            result = messagebox.showinfo(
                                "Đơn hàng mới!",
                                f"Mã đơn: #{order_id}\nKhách hàng: {safe_customer_name}\nTổng tiền: {total_amount:,.0f} VNĐ"
                            )
                            root.destroy()
                        
                        msg_thread = threading.Thread(target=show_messagebox)
                        msg_thread.daemon = True
                        msg_thread.start()
                        print("Messagebox notification sent")
                        
                    except Exception as e2:
                        print(f"Messagebox also failed: {e2}")
                        print("Only sound notification available")
                
            except Exception as e:
                print(f"Error showing notification: {e}")
        
        # Run in separate thread
        notification_thread = threading.Thread(target=show_notification)
        notification_thread.daemon = True
        notification_thread.start()
    
    def auto_open_admin_panel(self):
        """Automatically open admin panel when new order arrives"""
        if not self.enabled:
            return
            
        def open_admin():
            try:
                # Mở trình duyệt với admin panel
                import webbrowser
                admin_url = 'http://localhost:5000/admin'
                webbrowser.open(admin_url)
                print(f"Admin panel opened: {admin_url}")
                
                # Thử phương pháp pyautogui nếu cần
                if self.auto_minimize:
                    time.sleep(1)
                    pyautogui.hotkey('ctrl', 't')
                    time.sleep(0.5)
                    pyautogui.write(admin_url)
                    pyautogui.press('enter')
                    
            except Exception as e:
                print(f"Error opening admin panel: {e}")
        
        admin_thread = threading.Thread(target=open_admin)
        admin_thread.daemon = True
        admin_thread.start()
    
    def quick_order_print(self, order_id):
        """Quick print order details"""
        if pyautogui is None:
            print("pyautogui is not installed, cannot use quick print feature")
            return
        try:
            # Open order confirmation page
            pyautogui.hotkey('ctrl', 't')
            time.sleep(0.5)
            # Use localhost instead of hardcoded IP
            pyautogui.write(f'http://localhost:5000/order_confirmation/{order_id}')
            pyautogui.press('enter')
            time.sleep(2)
            
            # Print dialog
            pyautogui.hotkey('ctrl', 'p')
            time.sleep(0.5)
            pyautogui.press('enter')
            
        except Exception as e:
            print(f"Error printing order: {e}")
    
    def emergency_stop(self):
        """Emergency stop all automation"""
        self.enabled = False
        if pyautogui is not None:
            pyautogui.moveTo(0, 0)  # Trigger failsafe
        
    def get_screen_info(self):
        """Get screen information"""
        if pyautogui is None:
            return None
        try:
            return {
                'width': pyautogui.size().width,
                'height': pyautogui.size().height,
                'position': pyautogui.position()
            }
        except:
            return None

# Initialize Automation Controller
automation_controller = AutomationController()

# Allowed file extensions for upload
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file, prefix=''):
    """Save one uploaded file with a unique timestamped name. Returns the saved filename, or None if no valid file was given."""
    if not file or file.filename == '' or not allowed_file(file.filename):
        return None
    filename = secure_filename(file.filename)
    timestamp = int(time.time())
    filename = f"{timestamp}_{prefix}{filename}"

    upload_path = os.path.join(app.root_path, app.config['UPLOAD_FOLDER'])
    os.makedirs(upload_path, exist_ok=True)
    file.save(os.path.join(upload_path, filename))
    return filename

def save_uploaded_files(files, prefix='detail_'):
    """Save multiple uploaded files, returning the list of saved filenames (skips invalid entries)."""
    return [name for name in (save_uploaded_file(f, prefix=prefix) for f in files) if name]

# Initialize extensions
db = SQLAlchemy(app)
Session(app)

def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login'))
        return view(*args, **kwargs)
    return wrapped

def admin_required_api(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        return view(*args, **kwargs)
    return wrapped

def admin_required_api_success(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 401
        return view(*args, **kwargs)
    return wrapped

def super_admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login'))
        if session.get('admin_role') != 'super_admin':
            flash('Chỉ Super Admin mới có quyền truy cập chức năng này', 'error')
            return redirect(url_for('admin_dashboard'))
        return view(*args, **kwargs)
    return wrapped

def manager_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if 'customer_logged_in' not in session:
            return redirect(url_for('customer_login'))
        if session.get('customer_role') != 'manager':
            flash('Bạn không có quyền truy cập chức năng này', 'error')
            return redirect(url_for('customer_home'))
        return view(*args, **kwargs)
    return wrapped

def ensure_column(table_name, column_name, add_column_sql):
    """Add a column to a MySQL table if it doesn't already exist (idempotent, safe to call on every startup)."""
    try:
        exists = db.session.execute(text(
            "SELECT COUNT(*) FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = :table_name AND COLUMN_NAME = :column_name"
        ), {'table_name': table_name, 'column_name': column_name}).scalar()

        if not exists:
            print(f"Adding {column_name} column to {table_name} table...")
            db.session.execute(text(add_column_sql))
            db.session.commit()
            print(f"{column_name} column added successfully")
    except Exception as e:
        db.session.rollback()
        print(f"{table_name}.{column_name} migration check failed: {e}")

def populate_default_basic_costs():
    try:
        rooms = Room.query.all()
        modified = False
        for r in rooms:
            if not r.basic_costs:
                costs = [
                    {"name": "Điện", "value": getattr(r, 'electricity_cost', None) or '3.6k/kWh', "icon": "bolt"},
                    {"name": "Nước", "value": getattr(r, 'water_cost', None) or '100.000 VNĐ/người', "icon": "droplet"},
                    {"name": "Dịch vụ", "value": getattr(r, 'service_cost', None) or '100.000 VNĐ/người', "icon": "file-invoice"}
                ]
                r.basic_costs = json.dumps(costs)
                modified = True
        if modified:
            db.session.commit()
            print("Populated default basic costs for existing rooms successfully")
    except Exception as e:
        print(f"Room basic costs data population skipped: {e}")

# Models
class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    price = db.Column(db.Float, nullable=False)
    stock = db.Column(db.Integer, nullable=False)
    image = db.Column(db.String(200), default='pngtree.png')
    category = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    images = db.relationship('ProductImage', backref='product', lazy=True, cascade="all, delete-orphan")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
    
    @property
    def image_url(self):
        """Get the full URL for the product image"""
        if self.image and self.image not in ('pngtree.png', 'default_placeholder.png'):
            return url_for('static', filename=f'images/{self.image}')
        return url_for('static', filename='images/default_placeholder.png')

class ProductImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    image = db.Column(db.String(200), nullable=False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
    
    @property
    def image_url(self):
        """Get the full URL for this product image"""
        return url_for('static', filename=f'images/{self.image}')

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_name = db.Column(db.String(100), nullable=False)
    customer_phone = db.Column(db.String(20), nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='pending')
    payment_method = db.Column(db.String(20), default='cash')
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    items = db.relationship('OrderItem', backref='order', lazy=True)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class OrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Float, nullable=False)
    product = db.relationship('Product', backref='order_items')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default='admin', nullable=False)  # 'super_admin' or 'admin'

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    phone = db.Column(db.String(20), nullable=False)
    role = db.Column(db.String(20), default='customer', nullable=False)  # 'customer' or 'manager'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(50), nullable=False)
    message = db.Column(db.Text, nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text, nullable=False)
    price_per_hour = db.Column(db.Float, nullable=False)
    price_unit = db.Column(db.String(50), default='giờ')
    capacity = db.Column(db.Integer, nullable=False)
    image = db.Column(db.String(200), default='pngtree.png')
    amenities = db.Column(db.Text)  # JSON string of amenities
    available = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    electricity_cost = db.Column(db.String(100), default='3.6k/kWh')
    water_cost = db.Column(db.String(100), default='100.000 VNĐ/người')
    service_cost = db.Column(db.String(100), default='100.000 VNĐ/người')
    basic_costs = db.Column(db.Text)

    images = db.relationship('RoomImage', backref='room', lazy=True, cascade="all, delete-orphan")

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
    
    @property
    def image_url(self):
        """Get the full URL for the room image"""
        if self.image and self.image not in ('pngtree.png', 'default_placeholder.png'):
            return url_for('static', filename=f'images/{self.image}')
        return url_for('static', filename='images/default_placeholder.png')
    
    @property
    def amenities_list(self):
        """Get amenities as list"""
        if self.amenities:
            return json.loads(self.amenities)
        return []

    @property
    def basic_costs_list(self):
        """Get basic costs as list of dicts"""
        if self.basic_costs:
            try:
                return json.loads(self.basic_costs)
            except Exception:
                pass
        return []

class RoomImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('room.id'), nullable=False)
    image = db.Column(db.String(200), nullable=False)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
    
    @property
    def image_url(self):
        """Get the full URL for this room image"""
        return url_for('static', filename=f'images/{self.image}')

class RoomBooking(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('room.id'), nullable=False)
    customer_name = db.Column(db.String(100), nullable=False)
    customer_phone = db.Column(db.String(20), nullable=False)
    booking_date = db.Column(db.Date, nullable=False)
    start_time = db.Column(db.Time, nullable=False)
    end_time = db.Column(db.Time, nullable=False)
    total_hours = db.Column(db.Float, nullable=False)
    total_price = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, confirmed, cancelled
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    room = db.relationship('Room', backref='bookings')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

class DailyMenuItem(db.Model):
    """A dish a manager posts as available today. Ordered directly (like a room),
    not through the shopping cart."""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text)
    price = db.Column(db.Float, nullable=False)
    image = db.Column(db.String(200), default='pngtree.png')
    available = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by_id = db.Column(db.Integer, db.ForeignKey('customer.id'))

    created_by = db.relationship('Customer')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    @property
    def image_url(self):
        if self.image and self.image not in ('pngtree.png', 'default_placeholder.png'):
            return url_for('static', filename=f'images/{self.image}')
        return url_for('static', filename='images/default_placeholder.png')

class DailyMenuOrder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('daily_menu_item.id'), nullable=False)
    customer_name = db.Column(db.String(100), nullable=False)
    customer_phone = db.Column(db.String(20), nullable=False)
    quantity = db.Column(db.Integer, default=1)
    notes = db.Column(db.Text)
    status = db.Column(db.String(20), default='pending')  # pending, completed, cancelled
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    item = db.relationship('DailyMenuItem', backref='orders')

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

def create_notification(notification_type, message):
    try:
        notification = Notification(type=notification_type, message=message)
        db.session.add(notification)
        db.session.commit()
        return notification
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        print(f"Error creating notification: {e}")
        return None

# Routes
@app.route('/')
def index():
    return redirect(url_for('customer_home'))

@app.route('/customer')
def customer_home():
    products = Product.query.all()  # Hiển thị tất cả sản phẩm kể cả hết hàng
    daily_items = DailyMenuItem.query.filter_by(available=True).order_by(DailyMenuItem.created_at.desc()).all()
    return render_template('index.html', products=products, daily_items=daily_items)

@app.route('/admin/manage_products')
@admin_required
def manage_products():
    # Get search and filter parameters
    search = request.args.get('search', '')
    stock_filter = request.args.get('stock_filter', '')
    
    # Start with all products
    products = Product.query
    
    # Apply search filter
    if search:
        products = products.filter(Product.name.contains(search))
    
    # Apply stock filter
    if stock_filter == 'in_stock':
        products = products.filter(Product.stock > 0)
    elif stock_filter == 'out_of_stock':
        products = products.filter(Product.stock == 0)
    elif stock_filter == 'low_stock':
        products = products.filter(Product.stock > 0, Product.stock <= 10)
    
    # Get filtered products
    products = products.all()
    
    return render_template('manage_products.html', products=products)

@app.route('/products')
def products():
    # Get filter parameters
    search_query = request.args.get('search', '').strip()
    category_filter = request.args.get('category', '').strip()
    
    # Start with base query
    query = Product.query
    
    # Apply search filter
    if search_query:
        query = query.filter(Product.name.contains(search_query))
    
    # Apply category filter
    if category_filter:
        query = query.filter(Product.category == category_filter)
    
    # Get filtered products
    products = query.all()
    
    # Get all unique categories from database
    categories = db.session.query(Product.category).distinct().all()
    categories = [cat[0] for cat in categories if cat[0]]  # Remove None values
    
    # Get total products count
    total_products = Product.query.count()
    
    return render_template('products.html', products=products, categories=categories, total_products=total_products)

@app.route('/rooms')
def rooms():
    rooms = Room.query.filter_by(available=True).all()
    return render_template('rooms.html', rooms=rooms)

@app.route('/room/<int:room_id>')
def room_detail(room_id):
    room = Room.query.get_or_404(room_id)
    return render_template('room_detail.html', room=room)

@app.route('/book_room/<int:room_id>', methods=['POST'])
def book_room(room_id):
    room = Room.query.get_or_404(room_id)
    
    if not room.available:
        flash('Phòng này hiện không còn trống!', 'error')
        return redirect(url_for('room_detail', room_id=room_id))
    
    # Get form data
    customer_name = request.form.get('customer_name')
    customer_phone = request.form.get('customer_phone')
    booking_date_str = request.form.get('booking_date')
    start_time_str = request.form.get('start_time')
    notes = request.form.get('notes')
    
    # Validate required fields
    if not all([customer_name, customer_phone, booking_date_str, start_time_str]):
        flash('Vui lòng điền đầy đủ thông tin bắt buộc!', 'error')
        return redirect(url_for('room_detail', room_id=room_id))
    
    try:
        # Parse date and time
        from datetime import datetime, date, time, timedelta
        booking_date = datetime.strptime(booking_date_str, '%Y-%m-%d').date()
        try:
            start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
        except ValueError:
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
        # Fixed 1 hour booking duration
        start_datetime = datetime.combine(booking_date, start_time)
        end_datetime = start_datetime + timedelta(hours=1)
        end_time = end_datetime.time()
        
        # Fixed 1 hour booking
        total_hours = 1.0
        total_price = total_hours * room.price_per_hour

        # Reject overlapping bookings for the same room/date
        existing_bookings = RoomBooking.query.filter(
            RoomBooking.room_id == room.id,
            RoomBooking.booking_date == booking_date,
            RoomBooking.status != 'cancelled'
        ).all()
        for existing in existing_bookings:
            if existing.start_time < end_time and existing.end_time > start_time:
                flash('Khung giờ này đã có người đặt. Vui lòng chọn thời gian khác!', 'error')
                return redirect(url_for('room_detail', room_id=room_id))

        # Create booking
        booking = RoomBooking(
            room_id=room.id,
            customer_name=customer_name,
            customer_phone=customer_phone,
            booking_date=booking_date,
            start_time=start_time,
            end_time=end_time,
            total_hours=total_hours,
            total_price=total_price,
            notes=notes
        )
        
        db.session.add(booking)
        
        # Define end_time_str for notification
        end_time_str = end_time.strftime('%H:%M')
        # Create notification for admin
        notification_message = f"🏠 ĐẶT PHÒNG MỚI!\nKhách hàng: {customer_name}\nSĐT: {customer_phone}\nPhòng: {room.name}\nNgày: {booking_date_str}\nThời gian: {start_time_str} - {end_time_str}\nTổng: {total_price:,.0f} VNĐ"
        create_notification('room_booking', notification_message)
        
        db.session.commit()
        
        flash('Đặt phòng thành công! Chúng tôi sẽ liên hệ với bạn sớm.', 'success')
        # Redirect based on user role
        if 'admin_logged_in' in session:
            return redirect(url_for('admin_room_bookings'))
        else:
            return redirect(url_for('rooms'))
        
    except Exception as e:
        db.session.rollback()
        print(f"Booking error: {e}")
        flash('Có lỗi xảy ra khi đặt phòng. Vui lòng thử lại!', 'error')
        return redirect(url_for('room_detail', room_id=room_id))

@app.route('/order_daily_item/<int:item_id>', methods=['POST'])
def order_daily_item(item_id):
    item = DailyMenuItem.query.get_or_404(item_id)

    if not item.available:
        flash('Món này hiện không còn phục vụ', 'error')
        return redirect(url_for('customer_home'))

    customer_name = request.form.get('customer_name', '').strip() or 'Khách'
    customer_phone = request.form.get('customer_phone', '').strip()
    notes = request.form.get('notes', '').strip()

    try:
        quantity = max(1, int(request.form.get('quantity', 1)))
    except ValueError:
        quantity = 1

    if not customer_phone:
        flash('Vui lòng nhập số điện thoại để chúng tôi liên hệ', 'error')
        return redirect(url_for('customer_home'))

    order = DailyMenuOrder(
        item_id=item.id,
        customer_name=customer_name,
        customer_phone=customer_phone,
        quantity=quantity,
        notes=notes
    )
    db.session.add(order)
    create_notification('daily_menu_order', f"🍽️ Đặt món: {customer_name} đặt {quantity}x {item.name}")
    db.session.commit()

    flash(f'Đặt món "{item.name}" thành công! Chúng tôi sẽ liên hệ với bạn sớm.', 'success')
    return redirect(url_for('customer_home'))

# ---- Manager: daily menu management (logs in via /customer/login) ----
@app.route('/customer/daily-menu')
@manager_required
def daily_menu_manage():
    items = DailyMenuItem.query.order_by(DailyMenuItem.created_at.desc()).all()
    return render_template('daily_menu_manage.html', items=items)

@app.route('/customer/daily-menu/add', methods=['POST'])
@manager_required
def daily_menu_add():
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    price_raw = request.form.get('price', '').strip()

    if not name or not price_raw:
        flash('Vui lòng nhập tên món và giá', 'error')
        return redirect(url_for('daily_menu_manage'))

    try:
        price = float(price_raw)
    except ValueError:
        flash('Giá không hợp lệ', 'error')
        return redirect(url_for('daily_menu_manage'))

    image = save_uploaded_file(request.files.get('image')) or 'pngtree.png'

    item = DailyMenuItem(
        name=name,
        description=description,
        price=price,
        image=image,
        created_by_id=session.get('customer_id')
    )
    db.session.add(item)
    db.session.commit()
    flash(f'Đã thêm "{name}" vào thực đơn hôm nay', 'success')
    return redirect(url_for('daily_menu_manage'))

@app.route('/customer/daily-menu/<int:item_id>/toggle', methods=['POST'])
@manager_required
def daily_menu_toggle(item_id):
    item = DailyMenuItem.query.get_or_404(item_id)
    item.available = not item.available
    db.session.commit()
    return redirect(url_for('daily_menu_manage'))

@app.route('/customer/daily-menu/<int:item_id>/delete', methods=['POST'])
@manager_required
def daily_menu_delete(item_id):
    item = DailyMenuItem.query.get_or_404(item_id)
    name = item.name
    try:
        db.session.delete(item)
        db.session.commit()
        flash(f'Đã xóa "{name}" khỏi thực đơn', 'success')
    except Exception:
        db.session.rollback()
        flash(f'Không thể xóa "{name}" vì đã có đơn đặt liên quan — hãy tắt hiển thị thay vì xóa', 'error')
    return redirect(url_for('daily_menu_manage'))

# ---- Admin: view daily menu orders ----
@app.route('/admin/daily_menu_orders')
@admin_required
def admin_daily_menu_orders():
    orders = DailyMenuOrder.query.order_by(DailyMenuOrder.created_at.desc()).all()
    return render_template('admin_daily_menu_orders.html', orders=orders)

@app.route('/admin/daily_menu_order/<int:order_id>/update', methods=['POST'])
@admin_required
def admin_update_daily_menu_order(order_id):
    order = DailyMenuOrder.query.get_or_404(order_id)
    new_status = request.form.get('status')
    if new_status in ('pending', 'completed', 'cancelled'):
        order.status = new_status
        db.session.commit()
        flash('Đã cập nhật trạng thái đơn món ăn', 'success')
    return redirect(url_for('admin_daily_menu_orders'))

@app.route('/admin')
@app.route('/admin/login', methods=['GET'])
def admin_login():
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_authenticate():
    username = request.form.get('username')
    password = request.form.get('password')
    
    admin = Admin.query.filter_by(username=username).first()
    if admin and check_password_hash(admin.password, password):
        session['admin_logged_in'] = True
        session['admin_username'] = username
        session['admin_role'] = admin.role
        session.permanent = True  # Make session permanent
        return redirect(url_for('admin_dashboard'))
    else:
        flash('Thông tin đăng nhập không hợp lệ', 'error')
        return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    products = Product.query.all()
    orders = Order.query.order_by(Order.created_at.desc()).all()
    rooms = Room.query.all()
    room_bookings = RoomBooking.query.order_by(RoomBooking.created_at.desc()).all()
    notifications = Notification.query.order_by(Notification.created_at.desc()).limit(20).all()
    unread_notifications_count = Notification.query.filter_by(is_read=False).count()

    today = datetime.now().date()
    month_start = today.replace(day=1)

    completed_orders = [o for o in orders if o.status == 'completed']
    debt_orders = [o for o in orders if o.status not in ('completed', 'cancelled')]
    cancelled_orders = [o for o in orders if o.status == 'cancelled']
    today_completed = [o for o in completed_orders if o.created_at and o.created_at.date() == today]
    month_completed = [o for o in completed_orders if o.created_at and o.created_at.date() >= month_start]
    today_orders_count = len([o for o in orders if o.created_at and o.created_at.date() == today])

    top_products = db.session.query(
        Product.name, func.sum(OrderItem.quantity).label('total_qty')
    ).join(OrderItem, OrderItem.product_id == Product.id
    ).join(Order, Order.id == OrderItem.order_id
    ).filter(Order.status != 'cancelled'
    ).group_by(Product.id, Product.name
    ).order_by(func.sum(OrderItem.quantity).desc()
    ).limit(5).all()

    stats = {
        'total_revenue': sum(o.total_amount for o in completed_orders),
        'today_revenue': sum(o.total_amount for o in today_completed),
        'month_revenue': sum(o.total_amount for o in month_completed),
        'today_orders_count': today_orders_count,
        'completed_orders_count': len(completed_orders),
        'debt_orders_count': len(debt_orders),
        'total_debt': sum(o.total_amount for o in debt_orders),
        'cancelled_orders_count': len(cancelled_orders),
        'low_stock_count': len([p for p in products if 0 < p.stock <= 10]),
        'out_of_stock_count': len([p for p in products if p.stock == 0]),
        'available_rooms_count': len([r for r in rooms if r.available]),
        'pending_bookings_count': len([b for b in room_bookings if b.status == 'pending']),
        'top_products': top_products,
    }

    return render_template(
        'admin_dashboard.html',
        products=products,
        orders=orders,
        rooms=rooms,
        room_bookings=room_bookings,
        notifications=notifications,
        unread_notifications_count=unread_notifications_count,
        automation=automation_controller,
        laptop_speaker=laptop_speaker,
        stats=stats
    )

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    session.pop('admin_username', None)
    session.pop('admin_role', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/rooms')
@admin_required
def admin_rooms():
    rooms = Room.query.all()
    return render_template('admin_rooms.html', rooms=rooms)

@app.route('/admin/room/add', methods=['GET', 'POST'])
@admin_required
def admin_add_room():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        price_per_hour = float(request.form.get('price_per_hour'))
        capacity = int(request.form.get('capacity'))
        amenities_raw = request.form.get('amenities', '')
        amenities_list = [item.strip() for item in amenities_raw.split(',') if item.strip()]
        amenities = json.dumps(amenities_list)
        available = request.form.get('available') == 'on'

        image_url = save_uploaded_file(request.files.get('image'))

        cost_names = request.form.getlist('cost_names[]')
        cost_values = request.form.getlist('cost_values[]')
        cost_icons = request.form.getlist('cost_icons[]')

        costs_list = []
        for n, v, ic in zip(cost_names, cost_values, cost_icons):
            if n.strip() and v.strip():
                costs_list.append({
                    'name': n.strip(),
                    'value': v.strip(),
                    'icon': ic.strip()
                })
        basic_costs = json.dumps(costs_list)

        price_unit = request.form.get('price_unit', 'giờ')
        room = Room(  # type: ignore
            name=name,
            description=description,
            price_per_hour=price_per_hour,
            price_unit=price_unit,
            capacity=capacity,
            amenities=amenities,
            available=available,
            image=image_url,
            basic_costs=basic_costs
        )

        db.session.add(room)
        db.session.flush()  # Get room.id

        for filename in save_uploaded_files(request.files.getlist('images')):
            db.session.add(RoomImage(room_id=room.id, image=filename))

        db.session.commit()

        flash('Phòng đã được thêm thành công!', 'success')
        return redirect(url_for('admin_rooms'))
    
    return render_template('admin_add_room.html')

@app.route('/admin/room/<int:room_id>/toggle')
@admin_required
def admin_toggle_room(room_id):
    room = Room.query.get_or_404(room_id)
    room.available = not room.available
    db.session.commit()
    
    status = "mở" if room.available else "đóng"
    flash(f'Phòng {room.name} đã được {status}!', 'success')
    return redirect(url_for('admin_rooms'))

@app.route('/admin/room/<int:room_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_room(room_id):
    room = Room.query.get_or_404(room_id)
    
    if request.method == 'POST':
        room.image = save_uploaded_file(request.files.get('image')) or room.image

        for filename in save_uploaded_files(request.files.getlist('images')):
            db.session.add(RoomImage(room_id=room.id, image=filename))

        room.name = request.form.get('name')
        room.description = request.form.get('description')
        room.price_per_hour = float(request.form.get('price_per_hour'))
        room.price_unit = request.form.get('price_unit', 'giờ')
        room.capacity = int(request.form.get('capacity'))
        amenities_raw = request.form.get('amenities', '')
        amenities_list = [item.strip() for item in amenities_raw.split(',') if item.strip()]
        room.amenities = json.dumps(amenities_list)
        room.available = request.form.get('available') == 'on'
        cost_names = request.form.getlist('cost_names[]')
        cost_values = request.form.getlist('cost_values[]')
        cost_icons = request.form.getlist('cost_icons[]')

        costs_list = []
        for n, v, ic in zip(cost_names, cost_values, cost_icons):
            if n.strip() and v.strip():
                costs_list.append({
                    'name': n.strip(),
                    'value': v.strip(),
                    'icon': ic.strip()
                })
        room.basic_costs = json.dumps(costs_list)

        db.session.commit()

        flash('Phòng đã được cập nhật thành công!', 'success')
        return redirect(url_for('admin_rooms'))
    
    return render_template('admin_edit_room.html', room=room)

@app.route('/admin/room-image/<int:image_id>/delete', methods=['POST'])
@admin_required_api_success
def delete_room_image(image_id):
    img = RoomImage.query.get_or_404(image_id)
    try:
        # Delete file from disk
        image_path = os.path.join(app.root_path, 'static', 'images', img.image)
        if os.path.exists(image_path):
            os.remove(image_path)
    except Exception as e:
        print(f"Error deleting file from disk: {e}")
        
    db.session.delete(img)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/room/<int:room_id>/delete', methods=['POST'])
@admin_required
def admin_delete_room(room_id):
    room = Room.query.get_or_404(room_id)
    db.session.delete(room)
    db.session.commit()
    
    flash('Phòng đã được xóa thành công!', 'success')
    return redirect(url_for('admin_rooms'))


@app.route('/admin/room_bookings')
@admin_required
def admin_room_bookings():
    bookings = RoomBooking.query.order_by(RoomBooking.created_at.desc()).all()
    return render_template('admin_room_bookings.html', bookings=bookings)

@app.route('/admin/room_booking/<int:booking_id>/update', methods=['POST'])
@admin_required
def admin_update_room_booking(booking_id):
    booking = RoomBooking.query.get_or_404(booking_id)
    new_status = request.form.get('status')
    
    if new_status in ['pending', 'confirmed', 'cancelled']:
        booking.status = new_status
        db.session.commit()
        
        status_text = {
            'pending': 'Chờ xác nhận',
            'confirmed': 'Đã xác nhận', 
            'cancelled': 'Đã hủy'
        }[new_status]
        
        flash(f'Trạng thái đặt phòng đã được cập nhật: {status_text}', 'success')
    
    return redirect(url_for('admin_room_bookings'))

@app.route('/admin/product/add', methods=['GET', 'POST'])
@admin_required
def add_product():
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')
        price = float(request.form.get('price'))
        stock = int(request.form.get('stock'))
        category = request.form.get('category')

        image_url = save_uploaded_file(request.files.get('image'))

        product = Product(name=name, description=description, price=price, stock=stock, category=category, image=image_url)  # type: ignore[call-arg]
        db.session.add(product)
        db.session.flush()  # Get product.id before committing

        for filename in save_uploaded_files(request.files.getlist('images')):
            db.session.add(ProductImage(product_id=product.id, image=filename))

        db.session.commit()
        flash('Sản phẩm đã thêm thành công', 'success')
        return redirect(url_for('admin_dashboard'))
    
    return render_template('add_product.html')

@app.route('/admin/product/<int:product_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_product(product_id):
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        product.image = save_uploaded_file(request.files.get('image')) or product.image

        for filename in save_uploaded_files(request.files.getlist('images')):
            db.session.add(ProductImage(product_id=product.id, image=filename))

        # Update product details
        product.name = request.form.get('name')
        product.description = request.form.get('description')
        product.price = float(request.form.get('price'))
        product.stock = int(request.form.get('stock'))
        product.category = request.form.get('category')

        db.session.commit()
        flash('Sản phẩm đã cập nhật thành công', 'success')
        return redirect(url_for('manage_products'))
    
    return render_template('edit_product.html', product=product)

@app.route('/admin/product/<int:product_id>/delete', methods=['POST'])
@admin_required
def delete_product(product_id):
    product = Product.query.get_or_404(product_id)
    
    # Delete cover image if exists
    if product.image:
        try:
            image_path = os.path.join('static', 'images', product.image)
            if os.path.exists(image_path):
                os.remove(image_path)
        except Exception as e:
            print(f"Error deleting cover image: {e}")
            
    # Delete all detail images from disk
    for img in product.images:
        try:
            image_path = os.path.join('static', 'images', img.image)
            if os.path.exists(image_path):
                os.remove(image_path)
        except Exception as e:
            print(f"Error deleting detail image: {e}")
    
    # Delete product from database
    db.session.delete(product)
    db.session.commit()
    
    flash(f'Sản phẩm "{product.name}" đã được xóa thành công!', 'success')
    return redirect(url_for('manage_products'))

@app.route('/admin/product-image/<int:image_id>/delete', methods=['POST'])
@admin_required_api_success
def delete_product_image(image_id):
    img = ProductImage.query.get_or_404(image_id)
    try:
        # Delete file from disk
        image_path = os.path.join('static', 'images', img.image)
        if os.path.exists(image_path):
            os.remove(image_path)
    except Exception as e:
        print(f"Error deleting file from disk: {e}")
        
    db.session.delete(img)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/product/<int:product_id>')
def product_detail_route(product_id):
    product = Product.query.get_or_404(product_id)
    return render_template('product_detail.html', product=product)

@app.route('/cart')
def cart():
    cart_items = session.get('cart', [])
    products = []
    total = 0

    
    for item in cart_items:
        product = Product.query.get(item['product_id'])
        if product:
            products.append({
                'product': product,
                'quantity': item['quantity'],
                'subtotal': product.price * item['quantity']
            })
            total += product.price * item['quantity']
    
    final_total = total
    
    return render_template('cart.html', cart_items=products, products=products, subtotal=total, total=final_total)

@app.route('/add_to_cart/<int:product_id>', methods=['POST'])
def add_to_cart(product_id):
    product = Product.query.get_or_404(product_id)
    next_url = request.form.get('next') or request.referrer
    
    if product.stock <= 0:
        flash('Sản phẩm đã hết hàng', 'error')
        return redirect(next_url or url_for('customer_home'))
    
    # Get quantity from form, default to 1 if not provided
    quantity = int(request.form.get('quantity', 1))
    
    # Validate quantity
    if quantity < 1 or quantity > product.stock:
        flash('Số lượng không hợp lệ', 'error')
        return redirect(next_url or url_for('customer_home'))
    
    cart = session.get('cart', [])
    
    # Check if product already in cart
    for item in cart:
        if item['product_id'] == product_id:
            # Update quantity, but don't exceed stock
            new_quantity = item['quantity'] + quantity
            if new_quantity <= product.stock:
                item['quantity'] = new_quantity
            else:
                item['quantity'] = product.stock
                flash(f'Số lượng đã cập nhật tối đa ({product.stock})', 'warning')
            break
    else:
        cart.append({'product_id': product_id, 'quantity': quantity})
    
    session['cart'] = cart
    flash(f'{quantity} {"sản phẩm" if quantity == 1 else "sản phẩm"} đã thêm vào giỏ hàng', 'success')
    return redirect(next_url or url_for('customer_home'))

@app.route('/update_cart/<int:product_id>', methods=['POST'])
def update_cart(product_id):
    cart = session.get('cart', [])
    action = request.form.get('action')
    try:
        quantity = int(request.form.get('quantity', 1))
    except ValueError:
        quantity = 1

    product = Product.query.get(product_id)
    max_stock = product.stock if product else quantity

    # Find the item in cart
    for item in cart:
        if item['product_id'] == product_id:
            if action == 'increase':
                item['quantity'] += 1
            elif action == 'decrease' and item['quantity'] > 1:
                item['quantity'] -= 1
            else:
                item['quantity'] = quantity
            # Clamp to a valid range
            item['quantity'] = max(1, min(item['quantity'], max_stock)) if max_stock > 0 else 1
            break

    session['cart'] = cart
    return redirect(url_for('cart'))

@app.route('/remove_from_cart/<int:product_id>', methods=['POST', 'GET'])
def remove_from_cart(product_id):
    cart = session.get('cart', [])
    cart = [item for item in cart if item['product_id'] != product_id]
    session['cart'] = cart
    return redirect(url_for('cart'))

@app.route('/checkout', methods=['GET', 'POST'])
def checkout():
    cart_items = session.get('cart', [])
    if not cart_items:
        flash('Giỏ hàng của bạn đang trống', 'error')
        return redirect(url_for('customer_home'))
    
    # Handle POST from cart form
    if request.method == 'POST':
        # Get customer info from cart form
        customer_name = request.form.get('name', '').strip()
        customer_phone = request.form.get('phone', '').strip()
        
        # Store in session for checkout page
        session['customer_name'] = customer_name if customer_name else "Guest Customer"
        session['customer_phone'] = customer_phone if customer_phone else "Not provided"
        
        # Redirect to checkout page
        products = []
        total = 0
        
        for item in cart_items:
            product = Product.query.get(item['product_id'])
            if product:
                products.append({
                    'product': product,
                    'quantity': item['quantity'],
                    'subtotal': product.price * item['quantity']
                })
                total += product.price * item['quantity']
        
        final_total = total 
        return render_template('checkout.html', cart_items=products, subtotal=total, total=final_total)
    
    # Original GET logic
    products = []
    total = 0

    
    for item in cart_items:
        product = Product.query.get(item['product_id'])
        if product:
            products.append({
                'product': product,
                'quantity': item['quantity'],
                'subtotal': product.price * item['quantity']
            })
            total += product.price * item['quantity']
    
    final_total = total 
    return render_template('checkout.html', cart_items=products, subtotal=total,  total=final_total)

@app.route('/process_order', methods=['POST'])
def process_order():
    cart_items = session.get('cart', [])
    if not cart_items:
        flash('Giỏ hàng của bạn đang trống', 'error')
        return redirect(url_for('customer_home'))
    
    # Get form data
    customer_name = request.form.get('name', '').strip() or "Guest Customer"
    customer_phone = request.form.get('phone', '').strip() or "Not provided"
    payment_method = request.form.get('payment_method', 'cash')
    notes = request.form.get('notes', '').strip()

    phone_digits = re.sub(r"\D", "", customer_phone or "")
    if phone_digits:
        if len(phone_digits) < 9 or len(phone_digits) > 11:
            flash('Số điện thoại không hợp lệ (chỉ gồm 9-11 chữ số).', 'error')
            return redirect(url_for('checkout'))
        customer_phone = phone_digits

    # Verify stock is still available for every item before committing anything
    total = 0
    for item in cart_items:
        product = Product.query.get(item['product_id'])
        if not product:
            continue
        if item['quantity'] > product.stock:
            flash(f'Sản phẩm "{product.name}" không đủ hàng (còn {product.stock}). Vui lòng cập nhật giỏ hàng.', 'error')
            return redirect(url_for('cart'))
        total += product.price * item['quantity']

    final_total = total
    
    # Create order
    order = Order(
        customer_name=customer_name,
        customer_phone=customer_phone,
        total_amount=final_total,
        status='pending',
        payment_method=payment_method,
        notes=notes
    )
    db.session.add(order)
    db.session.flush()  # Get order ID
    
    # Add order items
    for item in cart_items:
        product = Product.query.get(item['product_id'])
        if product:
            order_item = OrderItem(
                order_id=order.id,
                product_id=product.id,
                quantity=item['quantity'],
                price=product.price
            )
            db.session.add(order_item)
            
            # Update stock
            product.stock -= item['quantity']
    
    db.session.commit()
    
    # Announce new order
    try:
        items_data = [{'product_id': item['product_id'], 'quantity': item['quantity']} for item in cart_items]
        laptop_speaker.announce_order(items_data, customer_name, notes=notes)
    except Exception as e:
        print(f"Error playing sound: {e}")
        
    # Clear cart
    session['cart'] = []

    create_notification('new_order', f"Đơn hàng mới #{order.id} - {customer_name} - {final_total:,.0f} VNĐ")
    
    # Create notification log
    try:
        log_entry = f"""
=====================================
🔔 ĐƠN HÀNG MỚI - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
=====================================
Mã đơn: #{order.id}
Khách hàng: {customer_name}
Số điện thoại: {customer_phone}
Phương thức thanh toán: {payment_method}
Tổng tiền: {final_total:,.0f} VNĐ
Sản phẩm: {len(cart_items)} loại
Ghi chú: {notes}
Trạng thái: {order.status}
=====================================
VUI LÒNG KIỂM TRA HỆ THỐNG ĐỂ XỬ LÝ ĐƠN HÀNG!
=====================================

"""
        
        with open('notification_log.txt', 'a', encoding='utf-8') as f:
            f.write(log_entry)
        
        print("✅ Notification log created: notification_log.txt")
        
    except Exception as e:
        print(f"Error creating notification log: {e}")
    
    flash('Đặt hàng thành công! Cảm ơn bạn đã mua hàng tại Coffee Vibes.', 'success')
    return redirect(url_for('order_confirmation', order_id=order.id))

@app.route('/admin/api/order/<int:order_id>')
@admin_required_api
def api_order_detail(order_id):
    """API endpoint to get order details for modal"""
    order = Order.query.get_or_404(order_id)
    
    # Get order items with product details
    items = []
    for item in order.items:
        product = Product.query.get(item.product_id)
        if product:
            items.append({
                'name': product.name,
                'quantity': item.quantity,
                'price': "{:,.0f}".format(item.price) + " VNĐ",
                'subtotal': "{:,.0f}".format(item.quantity * item.price) + " VNĐ"
            })
    
    # Format status badge
    status_badge = ""
    if order.status == 'pending':
        status_badge = '<span class="badge badge-warning">Chờ xử lý</span>'
    elif order.status == 'processing':
        status_badge = '<span class="badge badge-info">Đang xử lý</span>'
    elif order.status == 'completed':
        status_badge = '<span class="badge badge-success">Hoàn thành</span>'
    else:
        status_badge = '<span class="badge badge-danger">Hủy</span>'
    
    return jsonify({
        'id': order.id,
        'customer_name': order.customer_name,
        'customer_phone': order.customer_phone,
        'created_at': order.created_at.strftime('%d/%m/%Y %H:%M') if order.created_at else 'N/A',
        'status': order.status,
        'status_badge': status_badge,
        'total_amount': "{:,.0f}".format(order.total_amount) + " VNĐ",
        'notes': order.notes or '',
        'items': items
    })

@app.route('/admin/api/new_orders')
@admin_required_api
def api_new_orders():
    """Poll endpoint: orders placed after `since` (order id), for browser-side sound alerts"""
    since = request.args.get('since', 0, type=int)
    new_orders = Order.query.filter(Order.id > since).order_by(Order.id.asc()).all()

    result = []
    for order in new_orders:
        items = []
        for item in order.items:
            product = Product.query.get(item.product_id)
            items.append({
                'name': product.name if product else 'Sản phẩm đã xóa',
                'quantity': item.quantity
            })
        result.append({
            'id': order.id,
            'customer_name': order.customer_name,
            'customer_phone': order.customer_phone,
            'items': items,
            'notes': order.notes or '',
            'total_amount': "{:,.0f}".format(order.total_amount) + " VNĐ"
        })

    latest_id = new_orders[-1].id if new_orders else since
    return jsonify({'orders': result, 'latest_id': latest_id})

@app.route('/order_confirmation/<int:order_id>')
def order_confirmation(order_id):
    order = Order.query.get_or_404(order_id)
    # Get order items for display
    order_items = OrderItem.query.filter_by(order_id=order.id).all()
    
    # Calculate subtotal and total
    subtotal = order.total_amount  # Since we removed shipping fee, total_amount = subtotal
    total = order.total_amount
    
    return render_template('order_confirmation.html', order=order, order_items=order_items, subtotal=subtotal, total=total)

@app.route('/admin/order/<int:order_id>/update', methods=['POST'])
@admin_required
def update_order_status(order_id):
    order = Order.query.get_or_404(order_id)
    new_status = request.form.get('status')
    order.status = new_status
    db.session.commit()
    
    flash('Trạng thái đơn hàng đã cập nhật', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/debts')
@super_admin_required
def admin_debts():
    debt_orders = Order.query.filter(
        Order.status.notin_(['completed', 'cancelled'])
    ).order_by(Order.created_at.desc()).all()
    total_debt = sum(order.total_amount for order in debt_orders)
    return render_template('admin_debts.html', debt_orders=debt_orders, total_debt=total_debt)

@app.route('/admin/debt/<int:order_id>/pay', methods=['POST'])
@super_admin_required
def admin_debt_pay(order_id):
    order = Order.query.get_or_404(order_id)
    order.status = 'completed'
    db.session.commit()
    flash(f'Đã xác nhận {order.customer_name} trả hết nợ cho đơn #{order.id}', 'success')
    return redirect(url_for('admin_debts'))

@app.route('/admin/debts/bulk_pay', methods=['POST'])
@super_admin_required
def admin_debts_bulk_pay():
    order_ids = request.form.getlist('order_ids')
    if not order_ids:
        flash('Vui lòng chọn ít nhất một khoản nợ', 'error')
        return redirect(url_for('admin_debts'))

    count = Order.query.filter(Order.id.in_(order_ids)).update(
        {'status': 'completed'}, synchronize_session=False
    )
    db.session.commit()
    flash(f'Đã xóa {count} khoản nợ đã chọn', 'success')
    return redirect(url_for('admin_debts'))

@app.route('/admin/generate_qr', methods=['GET', 'POST'])
@admin_required
def generate_qr():
    if request.method == 'POST':
        # Get form data
        url = request.form.get('url', 'http://localhost:5000')
        size = int(request.form.get('size', 15))
        bg_color = request.form.get('bg_color', '#ffffff')
        fg_color = request.form.get('fg_color', '#000000')
        border = int(request.form.get('border', 1))
        
        # Generate QR code with custom parameters
        qr = qrcode.QRCode(
            version=1,
            box_size=size,
            border=border
        )
        qr.add_data(url)
        qr.make(fit=True)
        
        # Create image with custom colors
        img = qr.make_image(fill_color=fg_color, back_color=bg_color)
    else:
        # Default QR code for GET request
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        # Use localhost instead of hardcoded IP
        qr.add_data(f'http://localhost:5000/customer')
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="white", back_color="black")
    
    # Convert to base64 for display
    img_buffer = io.BytesIO()
    img.save(img_buffer)
    img_buffer.seek(0)
    img_base64 = base64.b64encode(img_buffer.getvalue()).decode()
    
    return render_template('qr_code.html', qr_code=img_base64)

@app.route('/admin/generate_bank_qr', methods=['GET', 'POST'])
@admin_required
def generate_bank_qr():
    qr_url = None
    if request.method == 'POST':
        bank_id = request.form.get('bank_id', 'VCB')
        account_no = request.form.get('account_no', '')
        account_name = request.form.get('account_name', '')
        amount = request.form.get('amount', '')
        add_info = request.form.get('add_info', '')
        template = request.form.get('template', 'print')
        
        import urllib.parse
        encoded_account_name = urllib.parse.quote(account_name)
        encoded_add_info = urllib.parse.quote(add_info)
        
        qr_url = f"https://img.vietqr.io/image/{bank_id}-{account_no}-{template}.jpg"
        params = []
        if amount:
            params.append(f"amount={amount}")
        if add_info:
            params.append(f"addInfo={encoded_add_info}")
        if account_name:
            params.append(f"accountName={encoded_account_name}")
            
        if params:
            qr_url += "?" + "&".join(params)
            
    return render_template('bank_qr.html', qr_url=qr_url)

@app.route('/customer/login')
def customer_login():
    if 'customer_logged_in' in session:
        return redirect(url_for('customer_home'))
    return render_template('customer_login.html')

@app.route('/customer/authenticate', methods=['POST'])
def customer_authenticate():
    username = request.form.get('username')
    password = request.form.get('password')
    
    customer = Customer.query.filter_by(username=username).first()
    if customer and check_password_hash(customer.password, password):
        session['customer_logged_in'] = True
        session['customer_id'] = customer.id
        session['customer_name'] = customer.full_name
        session['customer_role'] = customer.role
        flash('Đăng nhập thành công!', 'success')
        return redirect(url_for('customer_home'))
    else:
        flash('Thông tin đăng nhập không hợp lệ', 'error')
        return redirect(url_for('customer_login'))

@app.route('/customer/logout')
def customer_logout():
    session.pop('customer_logged_in', None)
    session.pop('customer_id', None)
    session.pop('customer_name', None)
    session.pop('customer_role', None)
    flash('Đăng xuất thành công!', 'success')
    return redirect(url_for('customer_home'))

@app.route('/customer/register')
def customer_register():
    return render_template('customer_register.html')

@app.route('/customer/create', methods=['POST'])
def customer_create():
    username = request.form.get('username')
    password = request.form.get('password')
    full_name = request.form.get('full_name') or username
    phone = request.form.get('phone') or ""
    
    # Check if username already exists
    if Customer.query.filter_by(username=username).first():
        flash('Tên đăng nhập đã tồn tại', 'error')
        return redirect(url_for('customer_register'))
    
    # Create new customer
    customer = Customer(
        username=username,
        password=generate_password_hash(password),
        full_name=full_name,
        phone=phone
    )
    db.session.add(customer)
    db.session.commit()
    
    flash('Đăng ký thành công!', 'success')
    return redirect(url_for('customer_home'))

# Excel Export Routes
@app.route('/admin/export_orders')
@super_admin_required
def export_orders():
    # Get today's date
    today = datetime.now().strftime('%Y-%m-%d')
    
    # Get orders for today
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = datetime.now().replace(hour=23, minute=59, second=59, microsecond=999999)
    
    orders = Order.query.filter(
        Order.created_at >= today_start,
        Order.created_at <= today_end
    ).order_by(Order.created_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orders_{today}"
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    
    # Create headers
    headers = ['Order ID', 'Customer Name', 'Phone', 'Total (VNĐ)', 'Payment Method', 'Status', 'Order Date']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row, order in enumerate(orders, 2):
        # Order ID
        ws.cell(row=row, column=1, value=f"#{order.id}").border = border
        
        # Customer Name
        ws.cell(row=row, column=2, value=order.customer_name).border = border
        
        # Phone
        ws.cell(row=row, column=3, value=order.customer_phone).border = border
        
        # Total Amount
        ws.cell(row=row, column=4, value=f"{order.total_amount:,.0f} VNĐ").border = border
        
        # Payment Method
        payment_display = {
            'cash': '💵 Cash',
            'bank_transfer': '🏦 Bank Transfer',
            'pending': '⏳ Pending'
        }.get(order.payment_method, order.payment_method)
        ws.cell(row=row, column=5, value=payment_display).border = border
        
        # Status
        status_display = {
            'pending': '⏳ Pending',
            'completed': '✅ Completed',
            'cancelled': '❌ Cancelled'
        }.get(order.status, order.status)
        ws.cell(row=row, column=6, value=status_display).border = border
        
        # Order Date
        ws.cell(row=row, column=7, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S')).border = border
    
    # Add summary section
    summary_row = len(orders) + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Total Orders: {len(orders)}")
    ws.cell(row=summary_row + 2, column=1, value=f"Total Revenue: {sum(order.total_amount for order in orders):,.0f} VNĐ")
    
    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create filename
    filename = f"orders_{today}.xlsx"
    
    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# Automation Control Routes
@app.route('/admin/automation_settings')
@super_admin_required
def automation_settings():
    screen_info = automation_controller.get_screen_info()
    return render_template('automation_settings.html',
                         automation=automation_controller,
                         laptop_speaker=laptop_speaker,
                         screen_info=screen_info)

@app.route('/admin/automation_toggle', methods=['POST'])
@super_admin_required
def automation_toggle():
    automation_controller.enabled = not automation_controller.enabled
    status = "bật" if automation_controller.enabled else "tắt"
    flash(f'Tự động hóa đã được {status}.', 'success')
    
    return redirect(url_for('automation_settings'))

@app.route('/admin/emergency_stop', methods=['POST'])
@super_admin_required
def emergency_stop():
    automation_controller.emergency_stop()
    flash('Đã dừng khẩn cấp tất cả tự động hóa!', 'warning')
    
    return redirect(url_for('automation_settings'))

@app.route('/admin/speaker_test', methods=['POST'])
@super_admin_required
def speaker_test():
    try:
        success = laptop_speaker.test_speaker()
        if success:
            flash('Kiểm tra loa laptop thành công!', 'success')
        else:
            flash('Kiểm tra loa laptop thất bại. Vui lòng kiểm tra cài đặt.', 'error')
    except Exception as e:
        flash(f'Lỗi kiểm tra loa: {e}', 'error')
    
    return redirect(url_for('automation_settings'))

@app.route('/admin/test_notification', methods=['POST'])
@super_admin_required
def test_notification():
    automation_controller.show_order_notification(999, "Khách test", 100000)
    flash('Đã gửi thông báo kiểm tra!', 'success')
    
    return redirect(url_for('automation_settings'))

# Laptop Speaker Control Routes
@app.route('/admin/speaker_toggle', methods=['POST'])
@super_admin_required
def speaker_toggle():
    enabled = laptop_speaker.toggle_enabled()
    status = "bật" if enabled else "tắt"
    flash(f'Thông báo loa laptop đã được {status}.', 'success')
    
    return redirect(url_for('automation_settings'))

@app.route('/admin/speaker_voice_settings', methods=['POST'])
@super_admin_required
def speaker_voice_settings():
    rate = request.form.get('voice_rate')
    volume = request.form.get('voice_volume')
    
    try:
        # Chuyển đổi thành số nếu có
        rate_value = int(rate) if rate else None
        volume_value = float(volume) if volume else None
        
        # Gọi set_voice_settings một lần với cả hai tham số
        laptop_speaker.set_voice_settings(rate=rate_value, volume=volume_value)
        
        flash('Cài đặt giọng nói đã được cập nhật!', 'success')
    except Exception as e:
        flash(f'Lỗi cập nhật cài đặt: {e}', 'error')
    
    return redirect(url_for('automation_settings'))

@app.route('/admin/toggle_tts_engine', methods=['POST'])
@super_admin_required
def toggle_tts_engine():
    try:
        # Toggle between gTTS and pyttsx3
        laptop_speaker.use_gtts = not laptop_speaker.use_gtts
        
        engine_type = "gTTS (Google Text-to-Speech)" if laptop_speaker.use_gtts else "pyttsx3 (Windows TTS)"
        flash(f'Đã chuyển sang sử dụng {engine_type}', 'success')
        
        # Reinitialize with new engine
        laptop_speaker.initialize_engine()
        
    except Exception as e:
        flash(f'Lỗi chuyển đổi engine TTS: {e}', 'error')
    
    return redirect(url_for('automation_settings'))

# Account management (Super Admin only)
@app.route('/admin/accounts')
@super_admin_required
def admin_accounts():
    admins = Admin.query.order_by(Admin.username).all()
    managers = Customer.query.filter_by(role='manager').order_by(Customer.username).all()
    customers = Customer.query.filter_by(role='customer').order_by(Customer.username).all()
    return render_template('admin_accounts.html', admins=admins, managers=managers, customers=customers)

@app.route('/admin/accounts/add_admin', methods=['POST'])
@super_admin_required
def admin_accounts_add_admin():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role', 'admin')

    if role not in ('admin', 'super_admin'):
        role = 'admin'

    if not username or not password:
        flash('Vui lòng nhập đầy đủ tên đăng nhập và mật khẩu', 'error')
        return redirect(url_for('admin_accounts'))

    if Admin.query.filter_by(username=username).first():
        flash('Tên đăng nhập admin đã tồn tại', 'error')
        return redirect(url_for('admin_accounts'))

    db.session.add(Admin(username=username, password=generate_password_hash(password), role=role))
    db.session.commit()
    flash(f'Đã tạo tài khoản {role} "{username}"', 'success')
    return redirect(url_for('admin_accounts'))

@app.route('/admin/accounts/<int:admin_id>/change_role', methods=['POST'])
@super_admin_required
def admin_accounts_change_role(admin_id):
    target = Admin.query.get_or_404(admin_id)
    new_role = request.form.get('role')
    if new_role not in ('admin', 'super_admin'):
        flash('Vai trò không hợp lệ', 'error')
        return redirect(url_for('admin_accounts'))

    if target.role == 'super_admin' and new_role == 'admin':
        remaining = Admin.query.filter(Admin.role == 'super_admin', Admin.id != target.id).count()
        if remaining == 0:
            flash('Không thể hạ quyền — đây là Super Admin cuối cùng', 'error')
            return redirect(url_for('admin_accounts'))

    target.role = new_role
    db.session.commit()
    flash(f'Đã đổi quyền của "{target.username}" thành {new_role}', 'success')
    return redirect(url_for('admin_accounts'))

@app.route('/admin/accounts/<int:admin_id>/delete_admin', methods=['POST'])
@super_admin_required
def admin_accounts_delete_admin(admin_id):
    target = Admin.query.get_or_404(admin_id)

    if target.username == session.get('admin_username'):
        flash('Không thể tự xóa tài khoản đang đăng nhập', 'error')
        return redirect(url_for('admin_accounts'))

    if target.role == 'super_admin':
        remaining = Admin.query.filter(Admin.role == 'super_admin', Admin.id != target.id).count()
        if remaining == 0:
            flash('Không thể xóa — đây là Super Admin cuối cùng', 'error')
            return redirect(url_for('admin_accounts'))

    db.session.delete(target)
    db.session.commit()
    flash(f'Đã xóa tài khoản "{target.username}"', 'success')
    return redirect(url_for('admin_accounts'))

@app.route('/admin/accounts/<int:customer_id>/toggle_manager', methods=['POST'])
@super_admin_required
def admin_accounts_toggle_manager(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    customer.role = 'customer' if customer.role == 'manager' else 'manager'
    db.session.commit()
    status = 'chỉ định làm Manager' if customer.role == 'manager' else 'gỡ quyền Manager'
    flash(f'Đã {status} cho "{customer.username}"', 'success')
    return redirect(url_for('admin_accounts'))

# Database initialization (runs on both local development and production Gunicorn import)
with app.app_context():
    try:
        db.create_all()
        ensure_column('product', 'image', "ALTER TABLE product ADD COLUMN image VARCHAR(200) DEFAULT 'placeholder.jpg'")
        ensure_column('room', 'basic_costs', "ALTER TABLE room ADD COLUMN basic_costs TEXT")
        ensure_column('room', 'price_unit', "ALTER TABLE room ADD COLUMN price_unit VARCHAR(50) DEFAULT 'giờ'")
        ensure_column('admin', 'role', "ALTER TABLE admin ADD COLUMN role VARCHAR(20) NOT NULL DEFAULT 'admin'")
        ensure_column('customer', 'role', "ALTER TABLE customer ADD COLUMN role VARCHAR(20) NOT NULL DEFAULT 'customer'")
        populate_default_basic_costs()

        # Create default admin if not exists
        admin = Admin.query.filter_by(username='admin').first()
        if not admin:
            admin = Admin(username='admin', password=generate_password_hash('admin123'), role='super_admin')
            db.session.add(admin)
            
            # Add some sample products
            products = [
                Product(name='Espresso', description='Cà phê đậm đậm nguyên chất từ hạt Arabica', price=45000, stock=50, category='coffee'),
                Product(name='Cappuccino', description='Cà phê với bọt sữa kem mịn', price=55000, stock=40, category='coffee'),
                Product(name='Latte', description='Cà phê sữa với lớp latte art đẹp mắt', price=60000, stock=35, category='coffee'),
                Product(name='Mocha', description='Cà phê kết hợp với chocolate đắng ngọt', price=65000, stock=30, category='coffee'),
                Product(name='Americano', description='Cà phê pha loãng với vị nguyên bản', price=50000, stock=45, category='coffee'),
                Product(name='Macchiato', description='Cà phê với chút sữa kem trên cùng', price=58000, stock=25, category='coffee'),
                Product(name='Flat White', description='Cà phê sữa với bọt mỏng', price=52000, stock=38, category='coffee'),
                Product(name='Cold Brew', description='Cà phê lạnh ngâm 24 giờ', price=48000, stock=42, category='coffee')
            ]
            
            for product in products:
                db.session.add(product)
            
            # Add sample rooms
            rooms = [
                Room(
                    name='Phòng VIP 1',
                    description='Phòng sang trọng với view đẹp, đầy đủ tiện nghi cao cấp, phù hợp cho họp nhóm hoặc nghỉ dưỡng.',
                    price_per_hour=150000,
                    capacity=4,
                    amenities=json.dumps(['WiFi', 'Điều hòa', 'TV 65 inch', 'Bàn làm việc', 'Ghế sofa', 'Máy pha cà phê', 'Mini bar']),
                    available=True
                ),
                Room(
                    name='Phòng Standard 2',
                    description='Phòng tiêu chuẩn với không gian ấm cúng, trang bị đầy đủ các tiện nghi cần thiết.',
                    price_per_hour=80000,
                    capacity=2,
                    amenities=json.dumps(['WiFi', 'Điều hòa', 'TV 43 inch', 'Bàn làm việc', 'Ghế văn phòng']),
                    available=True
                ),
                Room(
                    name='Phòng Family 3',
                    description='Phòng gia đình rộng rãi, có khu vực vui chơi nhỏ, phù hợp cho gia đình có trẻ em.',
                    price_per_hour=120000,
                    capacity=6,
                    amenities=json.dumps(['WiFi', 'Điều hòa', 'TV 55 inch', 'Bàn ăn', 'Ghế sofa', 'Khu vực vui chơi', 'Tủ lạnh']),
                    available=True
                ),
                Room(
                    name='Phòng Working 4',
                    description='Phòng làm việc chuyên nghiệp với thiết kế hiện đại, yên tĩnh, phù hợp cho làm việc nhóm.',
                    price_per_hour=100000,
                    capacity=3,
                    amenities=json.dumps(['WiFi', 'Điều hòa', 'Máy chiếu', 'Bàn họp', 'Ghế làm việc', 'Bảng trắng']),
                    available=True
                ),
                Room(
                    name='Phòng Relax 5',
                    description='Phòng thư giãn với không gian yên tĩnh, có thể nghe nhạc, đọc sách, thư giãn.',
                    price_per_hour=60000,
                    capacity=2,
                    amenities=json.dumps(['WiFi', 'Điều hòa', 'Loa nhạc', 'Ghế thư giãn', 'Sách', 'Trà']),
                    available=True
                )
            ]
            
            for room in rooms:
                db.session.add(room)
            db.session.commit()

        # One-time migration: re-hash any legacy plaintext passwords, and make
        # sure the founding 'admin' account is always super_admin (it may
        # have been created before the role column existed).
        legacy_dirty = False
        if admin.password.count('$') != 2:
            admin.password = generate_password_hash(admin.password)
            legacy_dirty = True
        if admin.role != 'super_admin':
            admin.role = 'super_admin'
            legacy_dirty = True
        for customer in Customer.query.all():
            if customer.password and customer.password.count('$') != 2:
                customer.password = generate_password_hash(customer.password)
                legacy_dirty = True
        if legacy_dirty:
            db.session.commit()
    except Exception as e:
        print(f"Error during database initialization: {e}")

if __name__ == '__main__':
    print("Registered Routes:")
    for rule in app.url_map.iter_rules():
        print(f"{rule.endpoint}: {rule}")
        
    app.run(debug=True, host='0.0.0.0', port=5000)
